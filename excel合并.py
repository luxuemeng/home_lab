
# coding: utf-8

# In[6]:


import pandas as pd
import os

def merge_data(id):
    base = r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\13"  #大文件夹,注意最后的/
    dir_names = os.listdir(base) # 打开大文件夹后的各个小文件夹名dir_names
    df_all = pd.DataFrame()
    for dir in dir_names:  #遍历每一个小文件夹
        file_path = os.path.join(base, dir)
        file_names = os.listdir(file_path)  #打开小文件夹后每一个excel文件的名称
        file_names1={'14议案投票概述_参.xlsx','19议员国会任职信息.xlsx'}
        for file in file_names1:  #遍历小文件夹里的每一个文件
            file_name = file_path+'/'+file
 
            if file.split('.')[0] == id:
                    df = pd.read_excel(file_name)
                   # df_all = df_all.append(df)
                    df_all = df_all.append(df,ignore_index=True)#索引重新排序
    return df_all

if __name__ == '__main__':
    
    base = r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\13"
    dir_names = os.listdir(base)
    #print(dir_names)
    ids = set()  # set()可以看做一个列表，这里面不包含重复的元素，不可以使用索引
    for dir in dir_names:
        file_path = os.path.join(base, dir)
        file_names = os.listdir(file_path)
        #print(file_names)
        file_names1={'14议案投票概述_参.xlsx','19议员国会任职信息.xlsx'}
        for file_name in file_names1:
            id = file_name.split('.')[0]  #id表示这一个股票的代码
            ids.add(id)  #把所有id放入ids中，这样不会有重复的id存在
            print(ids)
    for id in ids:
        df = merge_data(id)
        base =r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\15/"   #存放合并后的文件路径
        id_path = base + id + '.xlsx'
        df.to_excel(id_path,index=False)
#https://blog.csdn.net/qq_32649321/article/details/118416005


# In[ ]:


''' 
  * @author   remonl
  * @coding    utf-8
  * @date     2022/3/30 18:47
'''

import os
import pandas as pd


def get_files(path):
    fs = list()
    for root, dirs, files in os.walk(path):
        for file in files:
            fs.append(os.path.join(root, file))
    return fs


def merge():
    # 批量表所在文件夹路径
    files = get_files('../通用/')
    arr = list()
    for i in files:
        arr.append(pd.read_excel(i))
    # 读取文件行列数
    for j in arr:
        print(j.shape)
    # 目标文件的路径
    writer = pd.ExcelWriter('./通用.xlsx')
    pd.concat(arr).to_excel(writer, 'sheet1', index=False)
    writer.save()


if __name__ == '__main__':
    merge()


# In[ ]:


import pandas as pd
import os

def merge_data(id):
    base = 'path to Excel file/'  #大文件夹,注意最后的/
    dir_names = os.listdir(base) # 打开大文件夹后的各个小文件夹名dir_names
    df_all = pd.DataFrame()
    for dir in dir_names:  #遍历每一个小文件夹
        file_path = base+dir
        file_names = os.listdir(file_path)  #打开小文件夹后每一个excel文件的名称
        for file in file_names:  #遍历小文件夹里的每一个文件
            file_name = file_path+'/'+file
 
            if file.split('.')[0] == id:
                    df = pd.read_excel(file_name)
                   # df_all = df_all.append(df)
                    df_all = df_all.append(df,ignore_index=True)#索引重新排序
    return df_all

if __name__ == '__main__':
    
    base = 'path to excel file/'
    dir_names = os.listdir(base)
    ids = set()  # set()可以看做一个列表，这里面不包含重复的元素，不可以使用索引
    for dir in dir_names:
        file_path = base+dir
        file_names = os.listdir(file_path)
        for file_name in file_names:
            id = file_name.split('.')[0]  #id表示这一个股票的代码
            ids.add(id)  #把所有id放入ids中，这样不会有重复的id存在
    for id in ids:
        df = merge_data(id)
        base = 'path to save file/'  #存放合并后的文件路径
        id_path = base + id + '.xlsx'
        df.to_excel(id_path)


# In[ ]:


import os #用于获取文件路径
import xlrd #用于一次读取Excel中的整行数据
from openpyxl import load_workbook #用于写入数据

file_path="data"  # 文件所在文件夹

#1.获取路径下所有文件，并存入列表
pathss=[] # 存储文件夹内所有文件的路径（包括子目录内的文件）
for root, dirs, files in os.walk(file_path):
    path = [os.path.join(root, name) for name in files]
    pathss.extend(path)
    
#2.只提取出需要的Excel文件的路径
files_for_merge=[]
for i in pathss:
    if 'Tracker-sub' in i: #因文件夹内还有存储图片的Excel文件，需排除
        files_for_merge.append(i)
        
#3.读取各个Excel中的数据，并存入列表
data=[]
for i in files_for_merge:
    wb=xlrd.open_workbook(i) #按相应路径读取工作簿
    ws=wb.sheet_by_index(0) #选取工作表
    for j in range(10,ws.nrows):
        data.append(ws.row_values(j)) #读取整行数据，并存入列表  

#4.汇总数据到主Excel文件
wb_main=load_workbook(file_path+"/FM Reduction Activities Tracker-main.xlsx") #打开需要写入数据的文件
ws_main=wb_main['Raw Findings'] #选取需要写入数据的工作表
for row in range(3,len(data)+3):
    for col in range(1,18):
        ws_main.cell(row=row,column=col,value=data[row-3][col-1]) #写入数据
wb_main.save(file_path+"/FM Reduction Activities Tracker-main.xlsx") #保存数据
print("程序执行完成！")


# In[ ]:


import os #用于获取文件路径
import xlrd #用于一次读取Excel中的整行数据
from openpyxl import load_workbook #用于写入数据

file_path=r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\11"  # 文件所在文件夹

#1.获取路径下所有文件，并存入列表
pathss=[] # 存储文件夹内所有文件的路径（包括子目录内的文件）
for root, dirs, files in os.walk(file_path):
    path = [os.path.join(root, name) for name in files]
    pathss.extend(path)
    print(pathss)
    
#2.只提取出需要的Excel文件的路径
files_for_merge=[]
for i in pathss:
    if 'Tracker-sub' in i: #因文件夹内还有存储图片的Excel文件，需排除
        files_for_merge.append(i)
        
#3.读取各个Excel中的数据，并存入列表
data=[]
for i in files_for_merge:
    wb=xlrd.open_workbook(i) #按相应路径读取工作簿
    ws=wb.sheet_by_index(0) #选取工作表
    for j in range(10,ws.nrows):
        data.append(ws.row_values(j)) #读取整行数据，并存入列表  

#4.汇总数据到主Excel文件
wb_main=load_workbook(file_path+"/FM Reduction Activities Tracker-main.xlsx") #打开需要写入数据的文件
ws_main=wb_main['Raw Findings'] #选取需要写入数据的工作表
for row in range(3,len(data)+3):
    for col in range(1,18):
        ws_main.cell(row=row,column=col,value=data[row-3][col-1]) #写入数据
wb_main.save(file_path+"/FM Reduction Activities Tracker-main.xlsx") #保存数据
print("程序执行完成！")


# In[45]:


import os #用于获取文件路径
import xlrd #用于一次读取Excel中的整行数据
from openpyxl import load_workbook #用于写入数据

file_path=r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\11"  # 文件所在文件夹

#1.获取路径下所有文件，并存入列表
pathss=[] # 存储文件夹内所有文件的路径（包括子目录内的文件）
for root, dirs, files in os.walk(file_path):
    path = [os.path.join(root, name) for name in files]
    pathss.extend(path)
    print(pathss)


# In[49]:


import os
import pandas as pd
import numpy as np

dir = r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\13"#设置工作路径

#新建列表，存放文件名（可以忽略，但是为了做的过程能心里有数，先放上）
filename_excel = []

#新建列表，存放每个文件数据框（每一个excel读取后存放在数据框）
frames = []

for root, dirs, files in os.walk(dir):
    for file in files:
        #print(os.path.join(root,file))
        filename_excel.append(os.path.join(root,file))
        df = pd.read_excel(os.path.join(root,file)) #excel转换成DataFrame
        frames.append(df)
#打印文件名
print(filename_excel)   
 #合并所有数据
result = pd.concat(frames)    

#查看合并后的数据
result.head()
result.shape
base = r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\14"
#path=base+filename_excel +".xlsx"
result.to_excel(path)


# In[62]:


import os
import pandas as pd
import numpy as np

dir = r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\13"#设置工作路径

#新建列表，存放文件名（可以忽略，但是为了做的过程能心里有数，先放上）
filename_excel = []

#新建列表，存放每个文件数据框（每一个excel读取后存放在数据框）
frames = []

for root, dirs, files in os.walk(dir):
    for file in files:
        #print(file)
        #print(os.path.join(root,file))
        filename_excel.append(os.path.join(root,file))
        #print(filename_excel)
        df = pd.read_excel(os.path.join(root,file)) #excel转换成DataFrame
        frames.append(df)
#打印文件名
#print(filename_excel) 
print(files)
 #合并所有数据
result = pd.concat(frames)
dir = r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\14"
result.to_excel( r"D:\lxm\2022-11-电科云\爬虫数据校验\legislationOther\14\text.xlsx")

